'''
Parse the sign-ups table from excel into python objects
'''
import datetime
import json
from django.core.serializers.json import DjangoJSONEncoder
from django.core.management import BaseCommand
from django.db.transaction import atomic
from contextlib import contextmanager
from collections.abc import Generator
import os

from schedule.models import Event, EventInstance, Venue

from openpyxl import load_workbook

class DoRollback(Exception):
    pass

@contextmanager
def rollback_atomic() -> Generator[None, None, None]:
    try:
        with atomic():
            yield
            raise DoRollback()
    except:
        pass

class Command(BaseCommand):
    def handle(self, *args, **kwargs):

        workbook = load_workbook(filename="WSAF Excel Intranet.xlsx")
        ws = workbook['Signups by Submitter (8)']
        signups = ws.tables['WSAF24001.114']

        parent_events = {
            'Music Nights at Curiositea': 'Curiositea',
            'International Dance on The Windmill Hill Stage': 'Windmill Hill Stage',
            'Literature at Curiositea': 'Curiositea',
            'FAB Terrace': 'FAB Terrace',
            'Afternoon on The Hill': 'Windmill Hill Stage',
        }

        days = {
            'Saturday': datetime.date(2024, 6, 8),
            'Sunday': datetime.date(2024, 6, 9),
            'Monday': datetime.date(2024, 6, 10),
            'All Weekend': None,
        }

        events_to_sync = []
        for row in ws[signups.ref]:
            row_data = {}
            for index, cell in enumerate(row):
                row_data[signups.tableColumns[index].name] = cell.value
            # print(row_data)

            # Skip the header row
            if row_data['ID'] == 'ID':
                continue

            event = {}
            event['submission_id'] = row_data['ID']
            event['title'] = row_data['Name On A Programme']
            event['description'] = row_data['Description']
            event['estimated_runtime'] = row_data['Runtime']

            # try to convert Number Of Showings to int, if it fails, default to 1
            try:
                event['preferred_instances'] = int(row_data['Number Of Showings'] or 1)
            except ValueError:
                if str(row_data['Number Of Showings']).strip() in ['N/A', 'no', ""]:
                    event['preferred_instances'] = 1
                else:
                    # event['preferred_instances'] = 1
                    event['preferred_instances'] = input(f"Number Of Showings for {event['title']} is not a number: {row_data['Number Of Showings']}. Please enter a number: ")

            instances = []

            for i in range(1, 3):
                if row_data[f'Show {i} Venue'] is None:
                    continue
                instance_raw = {
                    'start_time': row_data[f'Show {i} Start Time'],
                    'end_time': row_data[f'Show {i} End Time'],
                    'date': days[row_data[f'Show {i} Day']],
                    'venue': row_data[f'Show {i} Venue']
                }
                # print(row_data[f'Show {i} Start Time'])

                if instance_raw['date'] is not None:
                    start_datetime = datetime.datetime.combine(instance_raw['date'], instance_raw['start_time'])
                    end_datetime = datetime.datetime.combine(instance_raw['date'], instance_raw['end_time'])
                if instance_raw['venue'] in parent_events:
                    instance_raw['parent_event'] = instance_raw['venue']
                    instance_raw['venue'] = parent_events[instance_raw['venue']]
                if instance_raw['venue'] == 'The Windmill Hill Stage':
                    instance_raw['venue'] = 'Windmill Hill Stage'
                instance = {
                    'start_time': start_datetime,
                    'end_time': end_datetime,
                    'venue': instance_raw['venue'],
                }
                if 'parent_event' in instance_raw:
                    instance['parent_event'] = instance_raw['parent_event']
                instances.append(instance)

            event['instances'] = instances

            events_to_sync.append(event)


        to_add_manually = []
        # with rollback_atomic():
        # print(events_to_sync)
        for event in events_to_sync:
            # print(event['submission_id'])
            matching_events = Event.objects.filter(submission_id=event['submission_id'])
            # print(matching_events.count())
            if matching_events.count() > 1:
                to_add_manually.append(event)
                print("MULTIPLE EVENTS FOR ", event['submission_id'])
                continue

            db_event: Event | None= matching_events.first()
            if db_event is None:
                print("NO DB EVENT FOR ", event['submission_id'])
                continue

            db_event.preferred_occurrences = event['preferred_instances']
            db_event.save()

            for instance in event['instances']:
                # print(instance)
                venue = Venue.objects.get(name=instance['venue'])
                db_instance = EventInstance(
                    event=db_event,
                    start=instance['start_time'],
                    end=instance['end_time'],
                    venue=venue
                )
                if 'parent_event' in instance:
                    if instance['parent_event'] == 'International Dance on The Windmill Hill Stage':
                        instance['parent_event'] = 'International Dance Showcase'
                    parent_event = Event.objects.get(title=instance['parent_event'])
                    parent_instance = parent_event.eventinstance_set.first()
                    db_instance.parent = parent_instance
                db_instance.save()

        # print(to_add_manually)
        for manual_add in to_add_manually:
            print(f'Add manually: {manual_add["submission_id"]} - {manual_add["title"]}')


# json.dump(events_to_sync, open('events_from_xslx.json', 'w', encoding='utf8'), indent=4, cls=DjangoJSONEncoder)
